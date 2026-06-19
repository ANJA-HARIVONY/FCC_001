#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exportación Excel de la lista de incidencias (página actual)."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

HEADER_FONT = Font(bold=True)

COLUMNS = (
    'ID',
    'Tipo cliente',
    'Cliente',
    'Dirección',
    'Asunto',
    'Estado',
    'Ref. Bitrix',
    'Usuario',
    'Fecha/Hora',
)

COLUMN_WIDTHS = {
    'A': 8,
    'B': 14,
    'C': 28,
    'D': 32,
    'E': 36,
    'F': 14,
    'G': 12,
    'H': 22,
    'I': 18,
}


def _incident_row(incident):
    from core.app import categoria_cliente_label

    client = incident.client
    operateur = incident.operateur
    fecha = ''
    if incident.date_heure:
        fecha = incident.date_heure.strftime('%d/%m/%Y %H:%M')
    return (
        incident.id,
        categoria_cliente_label(client.categoria) if client else '',
        client.nom if client else '',
        client.adresse if client else '',
        incident.intitule or '',
        incident.status or '',
        incident.ref_bitrix or '',
        operateur.nom if operateur else '',
        fecha,
    )


def build_incidents_list_workbook(incidents):
    """Genera un workbook Excel con solo la tabla de incidencias."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Incidencias'

    for col_idx, header in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header).font = HEADER_FONT

    for row_idx, incident in enumerate(incidents, start=2):
        for col_idx, value in enumerate(_incident_row(incident), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A2'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_export_filename():
    """Nombre de archivo: incidencias_{YYYYMMDD_HHMM}.xlsx"""
    return f"incidencias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
