#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exportación Excel de informes (etats)."""

import json
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

TYPE_LABELS_ES = {
    'summary': 'Estado General',
    'analysis': 'Evolución de Registros',
    'performance': 'Actividad por Usuario',
    'custom': 'Análisis Personalizado',
}

KPI_LABELS_ES = {
    'TOTAL_INCIDENTS': 'Total incidentes',
    'TOTAL_INCIDENTES': 'Total incidentes',
    'SOLUCIONADAS_DISTANCIA': 'Solucionadas (a distancia)',
    'SOLUCIONADAS': 'Solucionadas (a distancia)',
    'BITRIX': 'Bitrix (terreno)',
    'BITRIX_TERRENO': 'Bitrix (terreno)',
    'PENDIENTES': 'Pendientes',
    'TAUX_RESOLUTION': 'Tasa de resolución',
    'USUARIOS_ACTIVOS': 'Usuarios activos',
}

HEADER_FONT = Font(bold=True)


def parse_contenu_etat(etat):
    """Parse contenu_ia JSON del informe."""
    if not etat.contenu_ia:
        return None
    if isinstance(etat.contenu_ia, dict):
        return etat.contenu_ia
    try:
        return json.loads(etat.contenu_ia)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None


def _periode_label(etat):
    if etat.periode_debut and etat.periode_fin:
        return (
            f"{etat.periode_debut.strftime('%d/%m/%Y')} - "
            f"{etat.periode_fin.strftime('%d/%m/%Y')}"
        )
    if etat.periode_debut:
        return f"Desde el {etat.periode_debut.strftime('%d/%m/%Y')}"
    return 'No definido'


def _agence_label(contenu):
    if contenu and contenu.get('agence_label'):
        return contenu['agence_label']
    return 'Todas las agencias'


def _write_resumen_sheet(ws, etat, contenu):
    rows = [
        ('Campo', 'Valor'),
        ('Título', etat.titre),
        ('Tipo', TYPE_LABELS_ES.get(etat.type_etat, etat.type_etat)),
        ('Período', _periode_label(etat)),
        ('Agencia', _agence_label(contenu)),
        ('Estado', etat.statut),
        (
            'Generado el',
            etat.date_creation.strftime('%d/%m/%Y %H:%M') if etat.date_creation else '',
        ),
    ]
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = HEADER_FONT
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 48
    ws.freeze_panes = 'A2'


def _write_summary_datos(ws, contenu):
    row = 1
    ws.cell(row=row, column=1, value='Indicador').font = HEADER_FONT
    ws.cell(row=row, column=2, value='Valor').font = HEADER_FONT
    row += 1
    kpis = contenu.get('kpis_cles') or {}
    for key, value in kpis.items():
        label = KPI_LABELS_ES.get(key, key)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1
    for section_title, items_key in (
        ('Puntos positivos', 'points_positifs'),
        ('Puntos de atención', 'points_attention'),
        ('Recomendaciones', 'recommandations'),
    ):
        items = contenu.get(items_key) or []
        if not items:
            continue
        ws.cell(row=row, column=1, value=section_title).font = HEADER_FONT
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=str(item))
            row += 1
        row += 1
    if contenu.get('resume_executif'):
        ws.cell(row=row, column=1, value='Resumen ejecutivo').font = HEADER_FONT
        row += 1
        ws.cell(row=row, column=1, value=contenu['resume_executif'])
        row += 1
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 24
    ws.freeze_panes = 'A2'


def _write_analysis_datos(ws, contenu):
    headers = ['Fecha', 'Total', 'Solucionadas', 'Bitrix']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = HEADER_FONT
    evolution = contenu.get('evolution') or []
    for row_idx, point in enumerate(evolution, start=2):
        ws.cell(row=row_idx, column=1, value=point.get('label', ''))
        ws.cell(row=row_idx, column=2, value=point.get('count', 0))
        ws.cell(row=row_idx, column=3, value=point.get('solucionadas', 0))
        ws.cell(row=row_idx, column=4, value=point.get('bitrix', 0))
    if contenu.get('tendances_principales'):
        start = len(evolution) + 3
        ws.cell(row=start, column=1, value='Tendencias principales').font = HEADER_FONT
        for i, tendance in enumerate(contenu['tendances_principales'], start=start + 1):
            ws.cell(row=i, column=1, value=str(tendance))
    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 18
    ws.freeze_panes = 'A2'


def _write_performance_datos(ws, contenu):
    headers = ['Usuario', 'Total', 'Solucionadas', 'Bitrix', 'Eficacia (%)']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = HEADER_FONT
    rows = contenu.get('activite_utilisateurs') or []
    for row_idx, item in enumerate(rows, start=2):
        total = int(item.get('total') or 0)
        solucionadas = int(item.get('solucionadas') or 0)
        eficacia = round((solucionadas / total) * 100, 1) if total else 0
        ws.cell(row=row_idx, column=1, value=item.get('usuario', ''))
        ws.cell(row=row_idx, column=2, value=total)
        ws.cell(row=row_idx, column=3, value=solucionadas)
        ws.cell(row=row_idx, column=4, value=int(item.get('bitrix') or 0))
        ws.cell(row=row_idx, column=5, value=eficacia)
    kpis = contenu.get('kpis_cles') or {}
    if kpis:
        start = len(rows) + 3
        ws.cell(row=start, column=1, value='Indicadores agregados').font = HEADER_FONT
        row = start + 1
        for key, value in kpis.items():
            ws.cell(row=row, column=1, value=KPI_LABELS_ES.get(key, key))
            ws.cell(row=row, column=2, value=value)
            row += 1
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20
    ws.freeze_panes = 'A2'


def _write_custom_datos(ws, contenu):
    ws.cell(row=1, column=1, value='Análisis personalizado').font = HEADER_FONT
    text = contenu.get('analyse_personnalisee') or ''
    for row_idx, line in enumerate(text.splitlines() or [''], start=2):
        ws.cell(row=row_idx, column=1, value=line)
    ws.column_dimensions['A'].width = 80
    ws.freeze_panes = 'A2'


def build_etat_workbook(etat):
    """
    Genera un workbook Excel en memoria para un informe generado.

    Raises:
        ValueError: si el contenido no está disponible o es inválido.
    """
    contenu = parse_contenu_etat(etat)
    if contenu is None:
        raise ValueError('El contenido del informe no está disponible o es inválido.')

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = 'Resumen'
    _write_resumen_sheet(ws_resumen, etat, contenu)

    ws_datos = wb.create_sheet('Datos')
    type_etat = etat.type_etat or 'summary'
    if type_etat == 'summary':
        _write_summary_datos(ws_datos, contenu)
    elif type_etat == 'analysis':
        _write_analysis_datos(ws_datos, contenu)
    elif type_etat == 'performance':
        _write_performance_datos(ws_datos, contenu)
    elif type_etat == 'custom':
        _write_custom_datos(ws_datos, contenu)
    else:
        _write_summary_datos(ws_datos, contenu)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_export_filename(etat):
    """Nombre de archivo: informe_{id}_{YYYYMMDD_HHMM}.xlsx"""
    if etat.date_creation:
        ts = etat.date_creation.strftime('%Y%m%d_%H%M')
    else:
        ts = datetime.now().strftime('%Y%m%d_%H%M')
    return f'informe_{etat.id}_{ts}.xlsx'
