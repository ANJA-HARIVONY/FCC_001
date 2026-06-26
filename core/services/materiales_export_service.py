#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exportación Excel del informe de salidas de material."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from core.services.materiales_service import (
    MATERIAL_TIPO_LABELS,
    SALIDA_TIPO_LABELS,
    build_informe_rows,
)

HEADER_FONT = Font(bold=True)

COLUMNS = (
    'Fecha',
    'Técnico',
    'Cliente',
    'Material',
    'Tipo material',
    'Tipo salida',
    'Cantidad',
)

COLUMN_WIDTHS = {
    'A': 12,
    'B': 22,
    'C': 24,
    'D': 28,
    'E': 18,
    'F': 14,
    'G': 10,
}


def _linea_row_values(linea):
    salida = linea.salida
    material = linea.material
    fecha = salida.fecha.strftime('%d/%m/%Y') if salida and salida.fecha else ''
    tecnico = salida.tecnico.nom if salida and salida.tecnico else ''
    if salida and salida.client:
        cliente = salida.client.nom
    elif salida and salida.tipo_salida == 'uso_interno':
        cliente = 'Uso interno'
    else:
        cliente = 'Uso general'
    material_nombre = material.nombre if material else ''
    material_tipo = MATERIAL_TIPO_LABELS.get(material.tipo, material.tipo) if material else ''
    tipo_salida = SALIDA_TIPO_LABELS.get(salida.tipo_salida, salida.tipo_salida) if salida else ''
    return (
        fecha,
        tecnico,
        cliente,
        material_nombre,
        material_tipo,
        tipo_salida,
        linea.cantidad,
    )


def build_informe_workbook(date_from, date_to, agencia_id=None, filters=None, group_by=None):
    """Genera un workbook Excel con el detalle filtrado (sin subtotales)."""
    del group_by  # obsoleto
    lineas = build_informe_rows(date_from, date_to, agencia_id, filters=filters)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Informe salidas'

    for col_idx, header in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header).font = HEADER_FONT

    for row_idx, linea in enumerate(lineas, start=2):
        for col_idx, value in enumerate(_linea_row_values(linea), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A2'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_informe_export_filename():
    return f"informe_salidas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
