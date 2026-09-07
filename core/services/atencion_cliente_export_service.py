#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export Excel de la ficha Atención al cliente (retrasos)."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

HEADER_FONT = Font(bold=True)

COLUMNS = (
    'ID',
    'Cliente',
    'Ciudad',
    'Asunto',
    'Estado inicial',
    'Estado destino',
    'Fecha creación',
    'Fecha cambio',
    'Retraso (min)',
    'Ref. Bitrix',
    'Comentarios',
    'Último comentario',
    'Observaciones',
)

COLUMN_WIDTHS = {
    'A': 8,
    'B': 24,
    'C': 16,
    'D': 36,
    'E': 14,
    'F': 14,
    'G': 18,
    'H': 18,
    'I': 12,
    'J': 12,
    'K': 12,
    'L': 40,
    'M': 40,
}


def _fmt_dt(value):
    if not value:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%d/%m/%Y %H:%M')
    return str(value)


def build_status_delay_workbook(usuario, rows, filters, kpis):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Detalle incidencias'

    meta = [
        ('Usuario', usuario.nom if usuario else ''),
        ('Período desde', filters.get('date_from').strftime('%d/%m/%Y') if filters.get('date_from') else ''),
        ('Período hasta', filters.get('date_to').strftime('%d/%m/%Y') if filters.get('date_to') else ''),
        ('Estado destino', filters.get('estado_destino') or 'Todos'),
        ('Umbral (min)', filters.get('umbral') or 30),
        ('Total con retraso', kpis.get('total', 0)),
        ('Retraso medio', kpis.get('retraso_medio', 0)),
        ('Retraso máximo', kpis.get('retraso_max', 0)),
        ('Generado', datetime.now().strftime('%d/%m/%Y %H:%M')),
    ]
    for idx, (label, value) in enumerate(meta, start=1):
        ws.cell(row=idx, column=1, value=label).font = HEADER_FONT
        ws.cell(row=idx, column=2, value=value)

    header_row = len(meta) + 2
    for col_idx, header in enumerate(COLUMNS, start=1):
        ws.cell(row=header_row, column=col_idx, value=header).font = HEADER_FONT

    for row_idx, row in enumerate(rows, start=header_row + 1):
        values = (
            row.get('incident_id'),
            row.get('cliente'),
            row.get('ciudad') or row.get('cliente_ville') or '',
            row.get('asunto'),
            row.get('estado_inicial'),
            row.get('estado_destino'),
            _fmt_dt(row.get('fecha_creacion')),
            _fmt_dt(row.get('fecha_cambio')),
            row.get('retraso_min'),
            row.get('ref_bitrix') or '',
            row.get('comentarios_count', 0),
            row.get('ultimo_comentario') or '',
            row.get('observaciones') or '',
        )
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f'A{header_row + 1}'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_status_delay_export_filename(usuario):
    slug = (usuario.nom or 'usuario').replace(' ', '_')[:40]
    return f"ficha_retrasos_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
